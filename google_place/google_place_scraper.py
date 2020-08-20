import googlemaps
import pprint
import json
import xlsxwriter
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, fills
import os
import time

# Change working directory
os.chdir('C:\\Users\owner\Desktop')
CITY = 'London'


# function that re-formats the search result
# search result is a list of dictionary
def reformat_result(search_result):
    stored_results = []
    for result in search_result:
        # create an empty dictionary for each search result
        dict_result = {}

        address = result['formatted_address']

        coordinate_dict = result['geometry']['location']
        coordinate = ';'.join(str(x) for x in coordinate_dict.values())

        name = result['name']

        if 'price_level' in result:
            price_level = result['price_level']
        else:
            price_level = 'None'
        rating = result['rating']

        list = result['types']
        list.remove('point_of_interest')
        list.remove('establishment')
        for i in range(len(list)):
            list[i] = list[i].capitalize()
            list[i] += '(1)'

        types = ','.join([str(elem) for elem in list])

        rating_total = result['user_ratings_total']

        # add result to dictionary
        dict_result['name'] = name
        dict_result['syntax'] = '"{}"'.format(name)
        dict_result['relationship'] = types
        dict_result['address'] = address
        dict_result['geometry'] = coordinate
        dict_result['price level'] = price_level
        dict_result['rating'] = rating
        dict_result['total number of rating'] = rating_total

        # append dict to the list
        stored_results.append(dict_result)
    return stored_results


def to_excel(stored_results, text, workbook):
    # open the excel form and create a new worksheet

    worksheet = workbook.active
    worksheet.title = text

    row_num = worksheet.max_row
    # print (row_num)
    col_num = 1

    # populate other row

    # pprint.pprint(dict)
    for query in stored_results:
        result_values = query.values()

        # loop through each value in the value component
        for value in result_values:
            worksheet.cell(row=row_num + 1, column=col_num).value = value
            col_num += 1

        # make sure to get to next row and reset the column
        row_num += 1
        col_num = 1

    # # set the style for first row
    # # create a style
    # font = Font(bold=True, size=14)
    # fill = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
    # # workbook.add_named_style(title)
    # # apply the style
    # row_iter = worksheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=8)
    # for row in row_iter:
    #     for cell in row:
    #         cell.fill = fill
    #         cell.font = font

    # save the workbook
    workbook.save('map_data.xlsx')


def main():
    # Define API KEY and client
    API_KEY = os.environ.get('place_api_key')
    gmaps = googlemaps.Client(key=API_KEY)

    # read text file of restaurants names as a list
    with open('name.txt') as f:
        names = [line.rstrip() for line in f]

    # load excel sheet
    workbook = openpyxl.load_workbook('map_data.xlsx')
    worksheet = workbook.active
    # create row header
    row_headers = ['name', 'syntax', 'relationship', 'address', 'geometry', 'price_level', 'rating',
                   'total_number_of_rating']

    # populate the header row
    col_num = 1
    for header in row_headers:
        worksheet.cell(row=1, column=col_num).value = header.upper()
        col_num += 1

    workbook.save('map_data.xlsx')

    # read and search data from text file
    for name in names:
        search_text = f'{name} near {CITY}'

        places_result = gmaps.places(query=search_text, open_now=False, type='museum')
        search_results = places_result['results']
        time.sleep(3)

        if 'next_page_token' in places_result:
            places_result_2 = gmaps.places(query=search_text, page_token=places_result['next_page_token'])
            search_results_2 = places_result_2['results']
            search_results.extend(search_results_2)
        else:
            print(f'For {name}, no second page result.')

        if not places_result['results']:
            print(f'No {name} is found in the search.')
        else:
            query_result = reformat_result(search_results)
            to_excel(query_result, 'museum', workbook)


if __name__ == '__main__':
    main()
